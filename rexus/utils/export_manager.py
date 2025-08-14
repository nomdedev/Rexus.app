"""
Sistema de Exportación Estándar - Rexus.app v2.0.0

Sistema unificado de exportación de datos para todos los módulos del sistema.
Soporta exportación a Excel, CSV y PDF con configuraciones personalizables.
"""

import logging
import os
from datetime import datetime
from typing import Dict, List, Any, Optional

try:
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

import csv
from PyQt6.QtWidgets import QFileDialog, QWidget

from rexus.utils.message_system import show_success, show_error, show_warning


class ExportManager:
    """Gestor centralizado de exportación de datos."""

    def __init__(self):
        self.supported_formats = ['excel', 'csv', 'pdf']
        self.default_export_dir = self._get_default_export_directory()

    def _get_default_export_directory(self) -> str:
        """Obtiene el directorio por defecto para exportaciones."""
        try:
            # Crear directorio de exportaciones en Documents
            documents_dir = os.path.expanduser("~/Documents")
            export_dir = os.path.join(documents_dir, "Rexus_Exports")

            if not os.path.exists(export_dir):
                os.makedirs(export_dir, exist_ok=True)

            return export_dir
        except Exception as e:
            logging.error(f"Error creando directorio de exportación: {e}")
            return os.getcwd()

    def export_data(self,
                   data: List[Dict[str, Any]],
                   headers: List[str],
                   module_name: str,
                   export_format: str = 'excel',
                   filename: Optional[str] = None,
                   parent_widget: Optional[QWidget] = None) -> bool:
        """
        Exporta datos en el formato especificado.

        Args:
            data: Lista de diccionarios con los datos
            headers: Lista de encabezados de columnas
            module_name: Nombre del módulo que solicita la exportación
            export_format: Formato de exportación ('excel', 'csv', 'pdf')
            filename: Nombre personalizado del archivo (opcional)
            parent_widget: Widget padre para diálogos

        Returns:
            bool: True si la exportación fue exitosa
        """
        try:
            if export_format not in self.supported_formats:
                show_error(parent_widget, "Error", f"Formato {export_format} no soportado")
                return False

            if not data:
                show_warning(parent_widget, "Advertencia", "No hay datos para exportar")
                return False

            # Generar nombre de archivo si no se proporciona
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{module_name}_{timestamp}"

            # Ejecutar exportación según formato
            if export_format == 'excel':
                return self._export_to_excel(data,
headers,
                    filename,
                    parent_widget)
            elif export_format == 'csv':
                return self._export_to_csv(data,
headers,
                    filename,
                    parent_widget)
            elif export_format == 'pdf':
                return self._export_to_pdf(data,
headers,
                    filename,
                    parent_widget)

        except Exception as e:
            logging.error(f"Error en exportación: {e}")
            show_error(parent_widget, "Error", f"Error durante la exportación: {str(e)}")
            return False

    def _export_to_excel(self,
data: List[Dict],
        headers: List[str],
        filename: str,
        parent_widget: QWidget) -> bool:
        """Exporta datos a formato Excel."""
        if not OPENPYXL_AVAILABLE:
            show_error(parent_widget, "Error", "openpyxl no disponible. No se puede exportar a Excel")
            return False

        try:
            # Seleccionar ubicación del archivo
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Guardar archivo Excel",
                os.path.join(self.default_export_dir, f"{filename}.xlsx"),
                "Archivos Excel (*.xlsx);;Todos los archivos (*)"
            )

            if not file_path:
                return False  # Usuario canceló

            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Datos Exportados"

            # Configurar estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="366092")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Agregar encabezados
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Agregar datos
            for row_idx, record in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Ajustar ancho de columnas
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Guardar archivo
            wb.save(file_path)

            show_success(parent_widget, "Exportación Exitosa",
                        f"Datos exportados a Excel:\n{file_path}")
            return True

        except Exception as e:
            logging.error(f"Error exportando a Excel: {e}")
            show_error(parent_widget, "Error", f"Error exportando a Excel: {str(e)}")
            return False

    def _export_to_csv(self,
data: List[Dict],
        headers: List[str],
        filename: str,
        parent_widget: QWidget) -> bool:
        """Exporta datos a formato CSV."""
        try:
            # Seleccionar ubicación del archivo
            file_path, _ = QFileDialog.getSaveFileName(
                parent_widget,
                "Guardar archivo CSV",
                os.path.join(self.default_export_dir, f"{filename}.csv"),
                "Archivos CSV (*.csv);;Todos los archivos (*)"
            )

            if not file_path:
                return False  # Usuario canceló

            # Escribir archivo CSV
            with open(file_path,
'w',
                newline='',
                encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=headers)
                writer.writeheader()

                for record in data:
                    # Filtrar solo los campos que están en headers
                    filtered_record = {header: record.get(header, "") for header in headers}
                    writer.writerow(filtered_record)

            show_success(parent_widget, "Exportación Exitosa",
                        f"Datos exportados a CSV:\n{file_path}")
            return True

        except Exception as e:
            logging.error(f"Error exportando a CSV: {e}")
            show_error(parent_widget, "Error", f"Error exportando a CSV: {str(e)}")
            return False

    def _export_to_pdf(self,
data: List[Dict],
        headers: List[str],
        filename: str,
        parent_widget: QWidget) -> bool:
        """Exporta datos a formato PDF (implementación básica)."""
        try:
            # Por ahora, exportar como CSV y mostrar mensaje
            show_warning(
                parent_widget,
                "Funcionalidad PDF",
                "Exportación PDF en desarrollo.\nExportando como CSV en su lugar."
            )
            return self._export_to_csv(data, headers, filename, parent_widget)

        except Exception as e:
            logging.error(f"Error exportando a PDF: {e}")
            show_error(parent_widget, "Error", f"Error exportando a PDF: {str(e)}")
            return False

    def get_export_dialog_filters(self) -> str:
        """Retorna los filtros para diálogos de exportación."""
        filters = []
        if OPENPYXL_AVAILABLE:
            filters.append("Excel (*.xlsx)")
        filters.append("CSV (*.csv)")
        filters.append("Todos los archivos (*)")
        return ";;".join(filters)

    def validate_data_for_export(self,
data: List[Dict],
        headers: List[str]) -> tuple[bool,
        str]:
        """
        Valida que los datos sean apropiados para exportación.

        Returns:
            tuple: (es_valido, mensaje_error)
        """
        if not data:
            return False, "No hay datos para exportar"

        if not headers:
            return False, "No se han definido encabezados"

        if not isinstance(data, list):
            return False, "Los datos deben ser una lista"

        if not all(isinstance(record, dict) for record in data):
            return False, "Todos los registros deben ser diccionarios"

        return True, ""


class ModuleExportMixin:
    """
    Mixin para agregar funcionalidad de exportación a vistas de módulos.
    """

    def __init__(self):
        self.export_manager = ExportManager()

    def export_table_data(self, export_format: str = 'excel'):
        """
        Exporta los datos de la tabla principal del módulo.
        Debe ser sobrescrito por cada módulo.
        """
        # Esta es una implementación base que debe ser personalizada
        if hasattr(self, 'tabla_principal'):
            data = self._extract_table_data()
            headers = self._get_table_headers()
            module_name = self.__class__.__name__.replace('View', '').lower()

            return self.export_manager.export_data(
                data=data,
                headers=headers,
                module_name=module_name,
                export_format=export_format,
                parent_widget=self
            )
        else:
            show_error(self, "Error", "No se encontró tabla principal para exportar")
            return False

    def _extract_table_data(self) -> List[Dict[str, Any]]:
        """Extrae datos de la tabla principal."""
        data = []
        if hasattr(self, 'tabla_principal'):
            table = self.tabla_principal
            for row in range(table.rowCount()):
                record = {}
                for col in range(table.columnCount()):
                    header = table.horizontalHeaderItem(col)
                    item = table.item(row, col)

                    header_text = header.text() if header else f"Columna_{col}"
                    item_text = item.text() if item else ""

                    record[header_text] = item_text
                data.append(record)
        return data

    def _get_table_headers(self) -> List[str]:
        """Obtiene los encabezados de la tabla principal."""
        headers = []
        if hasattr(self, 'tabla_principal'):
            table = self.tabla_principal
            for col in range(table.columnCount()):
                header = table.horizontalHeaderItem(col)
                header_text = header.text() if header else f"Columna_{col}"
                headers.append(header_text)
        return headers

    def add_export_button(self, layout, button_text: str = "📄 Exportar"):
        """Agrega un botón de exportación al layout especificado."""
        try:
            from rexus.ui.components import RexusButton

            export_button = RexusButton(button_text, "secondary")
            export_button.clicked.connect(self.show_export_dialog)
            layout.addWidget(export_button)

        except Exception as e:
            logging.error(f"Error agregando botón de exportación: {e}")

    def show_export_dialog(self):
        """Muestra diálogo de selección de formato de exportación."""
        try:
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QRadioButton, QPushButton

            dialog = QDialog(self)
            dialog.setWindowTitle("Exportar Datos")
            dialog.setFixedSize(300, 200)

            layout = QVBoxLayout(dialog)

            # Opciones de formato
            self.excel_radio = QRadioButton("Excel (.xlsx)")
            self.csv_radio = QRadioButton("CSV (.csv)")
            self.pdf_radio = QRadioButton("PDF (.pdf)")

            # Seleccionar Excel por defecto
            self.excel_radio.setChecked(True)

            layout.addWidget(self.excel_radio)
            layout.addWidget(self.csv_radio)
            layout.addWidget(self.pdf_radio)

            # Botones
            button_layout = QHBoxLayout()

            export_btn = QPushButton("Exportar")
            export_btn.clicked.connect(lambda: self._execute_export_from_dialog(dialog))

            cancel_btn = QPushButton("Cancelar")
            cancel_btn.clicked.connect(dialog.reject)

            button_layout.addWidget(export_btn)
            button_layout.addWidget(cancel_btn)

            layout.addLayout(button_layout)

            dialog.exec()

        except Exception as e:
            logging.error(f"Error mostrando diálogo de exportación: {e}")
            show_error(self, "Error", f"Error mostrando diálogo: {str(e)}")

    def _execute_export_from_dialog(self, dialog):
        """Ejecuta la exportación según el formato seleccionado."""
        try:
            format_selected = 'excel'  # Default

            if hasattr(self, 'csv_radio') and self.csv_radio.isChecked():
                format_selected = 'csv'
            elif hasattr(self, 'pdf_radio') and self.pdf_radio.isChecked():
                format_selected = 'pdf'

            dialog.accept()
            self.export_table_data(format_selected)

        except Exception as e:
            logging.error(f"Error ejecutando exportación: {e}")
            show_error(self, "Error", f"Error en exportación: {str(e)}")


# Instancia global del gestor de exportación
export_manager = ExportManager()
